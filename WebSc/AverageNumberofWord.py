from openpyxl import load_workbook
import re

# Load the workbook
file_path_input = "Output.xlsx"
file_path_output = "Output Data Structure.xlsx"

# Load the input workbook
workbook_input = load_workbook(file_path_input)
sheet_input = workbook_input.active

# Load the output workbook
workbook_output = load_workbook(file_path_output)
sheet_output = workbook_output.active

# Function to count the average number of words per sentence
def count_average_words_per_sentence(content):
    # Split the content into sentences
    sentences = re.split(r'[.!?]+', content)
    
    # Remove empty sentences
    sentences = [sentence for sentence in sentences if sentence.strip()]
    
    # Count the total number of words
    total_words = sum(len(re.findall(r'\b\w+\b', sentence)) for sentence in sentences)
    
    # Count the total number of sentences
    total_sentences = len(sentences)
    
    # Calculate the average number of words per sentence
    average_words_per_sentence = total_words / total_sentences if total_sentences > 0 else 0
    
    return average_words_per_sentence

# Iterate over the rows starting from the second row in the input sheet
for row_index, row in enumerate(sheet_input.iter_rows(min_row=2, min_col=2, values_only=True), start=2):
    file_path = row[0]

    # Read the content of the file
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Calculate the average number of words per sentence
    average_words_per_sentence = count_average_words_per_sentence(content)
    
    # Fill the average number of words per sentence in the J column of the output sheet
    cell = sheet_output.cell(row=row_index, column=10)
    cell.value = average_words_per_sentence

# Save the output workbook
workbook_output.save(file_path_output)
