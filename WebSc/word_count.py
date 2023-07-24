from openpyxl import load_workbook
import os

# Function to count words in a file
def count_words_in_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
        words = content.split()
        return len(words)

# Load the file paths from the second column of "Output.xlsx"
input_file_path = "Output.xlsx"
output_file_path = "Output Data Structure.xlsx"
input_workbook = load_workbook(input_file_path)
output_workbook = load_workbook(output_file_path)

input_sheet = input_workbook.active
output_sheet = output_workbook.active

# Iterate over the rows starting from the second row
for row_index, row in enumerate(input_sheet.iter_rows(min_row=2, min_col=2)):
    file_path = row[0].value

    if file_path is not None and os.path.isfile(file_path):
        word_count = count_words_in_file(file_path)
        output_sheet.cell(row=row_index + 2, column=12, value=word_count)  # Fill the word count in the L column

# Save the updated output workbook
output_workbook.save(output_file_path)
