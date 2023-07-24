from openpyxl import load_workbook
import re

# Load the workbook
file_path_input = "Output.xlsx"
file_path_output = "Output Data Structure.xlsx"

# Load the input workbook
workbook_input = load_workbook(file_path_input)
sheet_input = workbook_input.active

# Load the output workbook
workbook_output = load_workbook(file_path_output)
sheet_output = workbook_output.active

# Function to count the number of syllables in a word
def count_syllables(word):
    # Remove non-alphabetic characters
    word = re.sub(r'[^a-zA-Z]', '', word)
    
    # Count the number of vowel sequences
    syllables = re.findall(r'[aeiouy]+', word, re.IGNORECASE)
    
    return len(syllables)

# Function to check if a word is complex
def is_complex(word):
    syllable_count = count_syllables(word)
    return syllable_count > 2

# Iterate over the rows starting from the second row in the input sheet
for row_index, row in enumerate(sheet_input.iter_rows(min_row=2, min_col=2, values_only=True), start=2):
    file_path = row[0]

    # Read the content of the file
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Split the content into words
    words = re.findall(r'\b\w+\b', content)

    # Count the complex words
    complex_word_count = sum(1 for word in words if is_complex(word))
    
    # Fill the complex word count in the K column of the output sheet
    cell = sheet_output.cell(row=row_index, column=11)
    cell.value = complex_word_count

# Save the output workbook
workbook_output.save(file_path_output)
