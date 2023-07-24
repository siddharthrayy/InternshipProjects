from openpyxl import load_workbook
import re

# Load the workbook
file_path_input = "Output.xlsx"
file_path_output = "Output Data Structure.xlsx"

# Load the input workbook
workbook_input = load_workbook(file_path_input)
sheet_input = workbook_input.active

# Load the output workbook
workbook_output = load_workbook(file_path_output)
sheet_output = workbook_output.active

# Function to check if a word is complex (contains more than two syllables)
def is_complex_word(word):
    # Count the number of syllables in the word
    syllable_count = len(re.findall(r'[aiouy]+e*|e(?!d$|ly).|[td]ed|le$', word, re.I))
    
    # Return True if the word has more than two syllables, False otherwise
    return syllable_count > 2

# Function to calculate the Fog Index
def calculate_fog_index(content):
    # Split the content into sentences
    sentences = re.split(r'(?<!\w\.\w.)(?<![A-Z][a-z]\.)(?<=\.|\?)\s', content)
    
    # Count the number of words
    word_count = len(re.findall(r'\b\w+\b', content))
    
    # Count the number of complex words
    complex_word_count = sum(is_complex_word(word) for word in re.findall(r'\b\w+\b', content))
    
    # Calculate the average sentence length
    average_sentence_length = word_count / len(sentences) if len(sentences) > 0 else 0
    
    # Calculate the percentage of complex words
    complex_word_percentage = (complex_word_count / word_count) * 100 if word_count > 0 else 0
    
    # Calculate the Fog Index
    fog_index = 0.4 * (average_sentence_length + complex_word_percentage)
    
    return fog_index

# Iterate over the rows starting from the second row in the input sheet
for row_index, row in enumerate(sheet_input.iter_rows(min_row=2, min_col=2, values_only=True), start=2):
    file_path = row[0]

    # Read the content of the file
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Calculate the Fog Index
    fog_index = calculate_fog_index(content)
    
    # Fill the Fog Index in the I column of the output sheet
    cell = sheet_output.cell(row=row_index, column=9)
    cell.value = fog_index

# Save the output workbook
workbook_output.save(file_path_output)
