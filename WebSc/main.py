import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

def scrape_medium(url, count):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    articles = soup.find_all('article')
    data = []

    for article in articles:
        heading = article.find('h1')
        if heading:
            data.append("Heading: " + heading.text.strip())

        paragraphs = article.find_all('p')
        body_text = "\n".join([p.text.strip() for p in paragraphs])
        if body_text:
            data.append("Body Text: " + body_text)

        data.append("")

    filename = f"{count}.txt"
    with open(filename, 'w', encoding='utf-8') as file:
        file.write('\n'.join(data))

    return filename

# Function to scrape Medium articles from URLs in an Excel file
def scrape_articles_from_excel(file_path):
    # Load the Excel file with data_only=True
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    # Create a new Excel file for storing the output filenames
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.append(["URL", "Output Filename"])

    # Read the URLs from the Excel file and scrape the articles
    count = 37  # Initial count value
    for row in sheet.iter_rows(min_row=2, values_only=True):
        url = row[1]  # Assuming the URLs are in the second column (B column)
        output_filename = scrape_medium(url, count)
        output_sheet.append([url, output_filename])
        count += 1

    # Save the output filenames to a new Excel file
    output_file_path = "Output.xlsx"
    output_workbook.save(output_file_path)

    # Save the updated count value
    workbook.save("Count.xlsx")


file_path = "Input.xlsx"
scrape_articles_from_excel(file_path)
 