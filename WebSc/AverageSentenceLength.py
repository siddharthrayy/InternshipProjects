from openpyxl import load_workbook
import re

# Load the workbook
file_path_input = "Output.xlsx"
file_path_output = "Output Data Structure.xlsx"

# Load the input workbook
workbook_input = load_workbook(file_path_input)
sheet_input = workbook_input.active

# Load the output workbook
workbook_output = load_workbook(file_path_output)
sheet_output = workbook_output.active

# Function to calculate the average sentence length
def calculate_average_sentence_length(content):
    # Split the content into sentences
    sentences = re.split(r'[.!?]+', content)
    
    # Remove empty sentences
    sentences = [sentence for sentence in sentences if sentence.strip()]
    
    # Count the number of words
    word_count = sum(len(re.findall(r'\b\w+\b', sentence)) for sentence in sentences)
    
    # Count the number of sentences
    sentence_count = len(sentences)
    
    # Calculate the average sentence length
    average_sentence_length = word_count / sentence_count if sentence_count > 0 else 0
    
    return average_sentence_length

# Iterate over the rows starting from the second row in the input sheet
for row_index, row in enumerate(sheet_input.iter_rows(min_row=2, min_col=2, values_only=True), start=2):
    file_path = row[0]

    # Read the content of the file
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Calculate the average sentence length
    average_sentence_length = calculate_average_sentence_length(content)
    
    # Fill the average sentence length in the G column of the output sheet
    cell = sheet_output.cell(row=row_index, column=7)
    cell.value = average_sentence_length

# Save the output workbook
workbook_output.save(file_path_output)
