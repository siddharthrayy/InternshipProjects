from openpyxl import load_workbook

# Load the workbook
file_path_input = "Output.xlsx"
file_path_output = "Output Data Structure.xlsx"

# Load the input workbook
workbook_input = load_workbook(file_path_input)
sheet_input = workbook_input.active

# Load the output workbook
workbook_output = load_workbook(file_path_output)
sheet_output = workbook_output.active

# Iterate over the rows starting from the second row in the input sheet
for row_index, row in enumerate(sheet_input.iter_rows(min_row=2, min_col=2, values_only=True), start=2):
    file_path = row[0]

    # Read the content of the file
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    # Calculate the average word length
    words = content.split()
    total_length = sum(len(word) for word in words)
    average_length = total_length / len(words) if len(words) > 0 else 0

    # Fill the average word length in the O column of the output sheet
    cell = sheet_output.cell(row=row_index, column=15)
    cell.value = average_length

# Save the output workbook
workbook_output.save(file_path_output)
