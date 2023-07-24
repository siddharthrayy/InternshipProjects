from openpyxl import load_workbook

# Load the workbook
file_path = "Output Data Structure.xlsx"
workbook = load_workbook(file_path)
sheet = workbook.active

# Iterate over the rows starting from the second row
for row in sheet.iter_rows(min_row=2):
    positive_score = row[2].value  # Third column (column C)
    negative_score = row[3].value  # Fourth column (column D)

    # Calculate the polarity score using the formula
    polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)

    # Fill the polarity score in the fifth column (column E)
    sheet.cell(row=row[0].row, column=5, value=polarity_score)

# Save the updated workbook
workbook.save(file_path)
