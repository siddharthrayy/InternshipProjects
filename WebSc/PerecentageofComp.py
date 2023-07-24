from openpyxl import load_workbook
import re

# Load the workbook
file_path_input = "Output.xlsx"
file_path_output = "Output Data Structure.xlsx"

# Load the input workbook
workbook_input = load_workbook(file_path_input)
sheet_input = workbook_input.active

# Load the output workbook
workbook_output = load_workbook(file_path_output)
sheet_output = workbook_output.active

# Function to check if a word is complex (contains more than two syllables)
def is_complex_word(word):
    # Count the number of syllables in the word
    syllable_count = len(re.findall(r'[aiouy]+e*|e(?!d$|ly).|[td]ed|le$', word, re.I))
    
    # Return True if the word has more than two syllables, False otherwise
    return syllable_count > 2

# Function to calculate the percentage of complex words
def calculate_complex_word_percentage(content):
    # Split the content into words
    words = re.findall(r'\b\w+\b', content)
    
    # Count the number of complex words
    complex_word_count = sum(is_complex_word(word) for word in words)
    
    # Count the total number of words
    total_word_count = len(words)
    
    # Calculate the percentage of complex words
    complex_word_percentage = (complex_word_count / total_word_count) * 100 if total_word_count > 0 else 0
    
    return complex_word_percentage

# Iterate over the rows starting from the second row in the input sheet
for row_index, row in enumerate(sheet_input.iter_rows(min_row=2, min_col=2, values_only=True), start=2):
    file_path = row[0]

    # Read the content of the file
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Calculate the percentage of complex words
    complex_word_percentage = calculate_complex_word_percentage(content)
    
    # Fill the complex word percentage in the H column of the output sheet
    cell = sheet_output.cell(row=row_index, column=8)
    cell.value = complex_word_percentage

# Save the output workbook
workbook_output.save(file_path_output)
