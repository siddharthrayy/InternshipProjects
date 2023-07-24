from openpyxl import load_workbook

# Load the workbook
file_path = "Output Data Structure.xlsx"
workbook = load_workbook(file_path)
sheet = workbook.active

# Iterate over the rows starting from the second row
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
    positive_score = row[2]  # Column C
    negative_score = row[3]  # Column D
    total_words = row[11]  # Column L

    subjectivity_score = (positive_score + negative_score) / (total_words + 0.000001)

    sheet.cell(row=row_index + 2, column=6, value=subjectivity_score)  # Fill the subjectivity score in column F

# Save the updated workbook
workbook.save(file_path)
