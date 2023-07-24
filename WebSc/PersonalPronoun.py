from openpyxl import load_workbook
import re

# Load the workbook
file_path_input = "Output.xlsx"
file_path_output = "Output Data Structure.xlsx"

# Load the input workbook
workbook_input = load_workbook(file_path_input)
sheet_input = workbook_input.active

# Load the output workbook
workbook_output = load_workbook(file_path_output)
sheet_output = workbook_output.active

# Define the personal pronouns
personal_pronouns = ["I", "you", "he", "she", "it", "we", "they", "me", "you", "him", "her", "us", "them", "myself",
                     "yourself", "himself", "herself", "itself", "ourselves", "yourselves", "themselves"]

# Iterate over the rows starting from the second row in the input sheet
for row_index, row in enumerate(sheet_input.iter_rows(min_row=2, min_col=2, values_only=True), start=2):
    file_path = row[0]

    # Read the content of the file
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    # Count the personal pronouns
    count = sum(1 for pronoun in personal_pronouns if re.search(r"\b" + pronoun + r"\b", content, re.IGNORECASE))

    # Fill the count in the N column of the output sheet
    cell = sheet_output.cell(row=row_index, column=14)
    cell.value = count

# Save the output workbook
workbook_output.save(file_path_output)
