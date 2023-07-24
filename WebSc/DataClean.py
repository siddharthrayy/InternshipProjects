from openpyxl import load_workbook
import os

def exclude_words_from_files(file_list, word_list_files):
    for file in file_list:
        file_directory = os.path.dirname(file)  # Extract the file directory

        with open(file, 'r', encoding='utf-8') as f:
            article = f.read()

        for word_list_file in word_list_files:
            with open(word_list_file, 'r', encoding='latin-1') as f:
                words_to_exclude = set(f.read().split())

            cleaned_article = ' '.join(word for word in article.split() if word.lower() not in words_to_exclude)

            cleaned_file_path = os.path.join(file_directory, os.path.basename(file))

            with open(cleaned_file_path, 'w', encoding='utf-8') as f:
                f.write(cleaned_article)

            # Update the file path in the output.xlsx with the cleaned file path
            update_file_path_in_excel(file, cleaned_file_path)


# Function to update the file path in the output.xlsx
def update_file_path_in_excel(old_file_path, new_file_path):
    file_path = "Output.xlsx"
    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        if row[0].value == old_file_path:
            cell = sheet.cell(row=row[0].row, column=2)  # Get the cell in the second column (column B)
            cell.value = new_file_path  # Update the cell value with the new file path
            break

    workbook.save(file_path)


# Load the file paths from the "B" column of Output.xlsx
file_path = "Output.xlsx"
workbook = load_workbook(file_path)
sheet = workbook.active

# Extract the file list from the specified column
file_list = []
for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
    if row[0] is not None:
        file_list.append(row[0])

# Provide the list of word list files to exclude
word_list_files = ['StopWords_Auditor.txt', 'StopWords_Currencies.txt', 'StopWords_DatesandNumbers.txt', 'StopWords_Generic.txt', 'StopWords_GenericLong.txt', 'StopWords_Geographic.txt', 'StopWords_Names.txt']

# Exclude words from files and save cleaned versions
exclude_words_from_files(file_list, word_list_files)
 