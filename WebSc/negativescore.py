from openpyxl import load_workbook

def count_matching_words(file1, file2):
    try:
        with open(file1, 'r', encoding='utf-8') as f1, open(file2, 'r', encoding='latin-1') as f2:
            article = f1.read()
            word_list = f2.read().split()

            matching_words = 0
            for word in word_list:
                if word.lower() in article.lower():
                    matching_words -= 1
 
            return matching_words * -1

    except FileNotFoundError:
        print("One or both files not found.")
    except IOError:
        print("Error reading file(s).")

# Load the file paths from the "B" column of Output.xlsx
input_file_path = "Output.xlsx"
output_file_path = "Output Data Structure.xlsx"

input_workbook = load_workbook(input_file_path, data_only=True)
input_sheet = input_workbook.active
file1_list = [row[0] for row in input_sheet.iter_rows(min_row=2, min_col=2, values_only=True)]

file2 = 'negative-words.txt'

output_workbook = load_workbook(output_file_path)
output_sheet = output_workbook.active

for row_index, file1 in enumerate(file1_list, start=2):
    count = count_matching_words(file1, file2)
    output_sheet.cell(row=row_index, column=4, value=count)

# Save the output Excel file
output_workbook.save(output_file_path)
